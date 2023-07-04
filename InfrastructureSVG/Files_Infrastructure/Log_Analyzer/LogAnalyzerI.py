import json
import locale
import os
import re
import socket
from inspect import currentframe, getframeinfo
from os import listdir
from pathlib import Path

import openpyxl
import yaml

from InfrastructureSVG.Files_Infrastructure.Log_Analyzer.input_validation.InputValidation import InputValidation
from InfrastructureSVG.Files_Infrastructure.Log_Analyzer.output.Output import Output
from InfrastructureSVG.Files_Infrastructure.Log_Analyzer.preprocessing.Preprocessing import Preprocessing
from InfrastructureSVG.Files_Infrastructure.Log_Analyzer.tracking.Logger import Logger
from InfrastructureSVG.Files_Infrastructure.Log_Analyzer.tracking.Push_notification import Notifications
from InfrastructureSVG.Files_Infrastructure.Log_Analyzer.validation.Validation import Validation


class LogAnalyzerI:
    def __init__(self):
        self.logger = Logger()
        self.input_validation = InputValidation()
        self.preprocessing = Preprocessing()
        self.validation = Validation()
        self.notifications = Notifications('rblumberg')
        self.output = Output()
        self.time_stats = {}
        self.validation_state = {}
        self.debug_flag = None
        self.flow_yaml_file_path = None
        self.log_file_path = None
        self.debug_test_lines_file_path = None
        self.inflated_output_file_path = None
        self.parsed_log_file_path = None
        self.log_pattern = None
        self.KB_to_MB = None
        self.reg_sequence_item_postfix = None
        self.compiled_regexes = None
        self.actions_types = None

    @staticmethod
    def add_args_to_global_configuration(global_configuration, args_dict):
        """
        Adds arguments that where received to global configuration
        Arguments:
            global_configuration: The global configuration object.
            args_dict: The arguments' dictionary.
        """
        for args_dic in args_dict:
            global_configuration[args_dic] = {}
            for key_index, key in enumerate(list(args_dict[args_dic].keys())):
                global_configuration[args_dic][key] = list(
                    args_dict[args_dic].values())[key_index]

    def init_analyzer(self, yaml_data):
        """
        Initialize the Log Analyzer
        Arguments:
            yaml_data: The YAML data object.
        """
        self.debug_flag = True
        self.KB_to_MB = 1 / 1024 / 1024
        self.reg_sequence_item_postfix = "_(.*)"
        if yaml_data['global_configuration'] is not None:
            self.validation.validation_state = {
                'global_configuration': yaml_data['global_configuration']}
        else:
            self.validation.validation_state = {'global_configuration': 'None'}

    def prepare_log_list(
            self,
            yaml_data,
            sir_folder_path,
            is_log_analyzer_test,
            under_or_supporting_test,
            set_filter=None):
        """
        Prepares a list of the logs' dictionary objects
        Arguments:
            yaml_data: The YAML data object.
            sir_folder_path: the sir folder path.
            is_log_analyzer_test: Weather the Log Analyzer is in Test Mode.
        Returns:
            log_list: The list of the logs' dictionary objects.
            flow_type_device_count: The number of rus.
        """
        flow_type_dict = {'CUCP': {'folder_name_prefix': 'cucp_(id)_SSH_Log'},
                          'CUUP': {'folder_name_prefix': 'cuup_(id)_SSH_Log'},
                          'DU_MAC': {'folder_name_prefix': 'du_(id)_SSH_Log'},
                          'DU_PHY': {'folder_name_prefix': 'du_(id)_SSH_Log'}}
        log_list = []
        flow_type_device_count = None
        frame_info = getframeinfo(currentframe())
        track_message = f"checking sequence_flows key in yaml_data"
        self.logger.info(frame_info, track_message)
        if 'sequence_flows' in yaml_data:
            frame_info = getframeinfo(currentframe())
            track_message = f"sequence_flows key found in yaml_data"
            self.logger.info(frame_info, track_message)
            frame_info = getframeinfo(currentframe())
            track_message = f"looping sequence_flow_name in yaml_data['sequence_flows']"
            self.logger.info(frame_info, track_message)
            for sequence_flow_name in yaml_data['sequence_flows']:
                sequence_flow = yaml_data['sequence_flows'][sequence_flow_name]
                frame_info = getframeinfo(currentframe())
                track_message = f"checking 'flow_configuration' in sequence_flow"
                self.logger.info(frame_info, track_message)
                if 'flow_configuration' in sequence_flow:
                    frame_info = getframeinfo(currentframe())
                    track_message = f"'flow_configuration' found in sequence_flow"
                    self.logger.info(frame_info, track_message)
                    sequence_flow_configuration = sequence_flow['flow_configuration']
                    frame_info = getframeinfo(currentframe())
                    track_message = f"checking 'flow_type' in sequence_flow_configuration"
                    self.logger.info(frame_info, track_message)
                    if 'flow_type' in sequence_flow_configuration:
                        frame_info = getframeinfo(currentframe())
                        track_message = f"'flow_type' found in sequence_flow_configuration"
                        self.logger.info(frame_info, track_message)
                        flow_type = sequence_flow_configuration['flow_type']
                        frame_info = getframeinfo(currentframe())
                        track_message = f"getting device list"
                        self.logger.info(frame_info, track_message)
                        device_dict_list = self.get_device_dict_list(sir_folder_path, flow_type_dict[flow_type]['folder_name_prefix'], under_or_supporting_test, set_filter)
                        frame_info = getframeinfo(currentframe())
                        track_message = f"device_dict_list: {device_dict_list}"
                        self.logger.info(frame_info, track_message)
                        for device_dict in device_dict_list:
                            last_log_file_path = self.get_last_log_from_folder(device_dict['device_id'],
                                                                               device_dict['device_name'],
                                                                               device_dict['device_dir_path'],
                                                                               flow_type,
                                                                               is_log_analyzer_test)
                            if last_log_file_path == 'multiple_logs':
                                frame_info = getframeinfo(currentframe())
                                track_message = f'muiltiple logs of the flow_type {flow_type} were found in the same folder.'
                                self.logger.error(frame_info, track_message)
                                break
                            if last_log_file_path == 'incompatible_folder_and_file_device_name':
                                frame_info = getframeinfo(currentframe())
                                track_message = f'The device name from the log file is incompatible to the folder device name.'
                                self.logger.error(frame_info, track_message)
                                break
                            if not last_log_file_path:
                                frame_info = getframeinfo(currentframe())
                                track_message = f'The last log path for one or more of the folders of the flow_type {flow_type} was not found.'
                                self.logger.error(frame_info, track_message)
                                break
                            flow_type_device_count = len(device_dict_list)
                            # regex_type = self.preprocessing.check_regex_type(flow_type, last_log_file_path)
                            regex_type = 'regular'
                            frame_info = getframeinfo(currentframe())
                            track_message = f'regex_type: {regex_type}'
                            self.logger.info(frame_info, track_message)
                            log_start_timestamp, start_message_id, log_end_timestamp, end_message_id = self.preprocessing.get_log_start_end_timestamps(regex_type, last_log_file_path)
                            log_dict = {
                                'log_type': flow_type,
                                'regex_type': regex_type,
                                'log_file_path': last_log_file_path,
                                'gnb_id': device_dict['gnb_id'],
                                'gnb_name': device_dict['gnb_name'],
                                'device_id': device_dict['device_id'],
                                'device_name': device_dict['device_name'],
                                'device_dir_path': device_dict['device_dir_path'],
                                'flow_type_device_count': flow_type_device_count,
                                'log_start_timestamp': log_start_timestamp,
                                'log_end_timestamp': log_end_timestamp,
                                'start_message_id': start_message_id,
                                'end_message_id': end_message_id
                            }
                            log_list.append(log_dict)
                    else:
                        frame_info = getframeinfo(currentframe())
                        track_message = 'A sequence flow has been specified but no flow_type has been specified in flow configuration'
                        self.logger.error(frame_info, track_message)
                        break
                else:
                    frame_info = getframeinfo(currentframe())
                    track_message = 'A sequence flow has been specified but no flow configuration has been specified'
                    self.logger.error(frame_info, track_message)
                    break
        return log_list, flow_type_device_count

    def get_device_dict_list(self, sir_folder_path, folder_name_prefix, under_or_supporting_test, set_filter=None):
        device_dict_list = []
        gnb_folder_index = 0
        frame_info = getframeinfo(currentframe())
        track_message = f"looping gnb_dir"
        self.logger.info(frame_info, track_message)
        for gnb_dir in listdir(sir_folder_path):
            if 'gnb' in gnb_dir:
                if under_or_supporting_test in gnb_dir:
                    gnb_folder_name_prefix = 'gnb_(id)'
                    gnb_folder_name_prefix_regex = re.sub('.*\((id)\).*', '.*_([0-9]+)_.*', gnb_folder_name_prefix)
                    gnb_id_search = re.search(gnb_folder_name_prefix_regex, gnb_dir)
                    frame_info = getframeinfo(currentframe())
                    track_message = f"gnb_dir: {gnb_dir}"
                    self.logger.info(frame_info, track_message)
                    frame_info = getframeinfo(currentframe())
                    track_message = f"extracting gnb_id"
                    self.logger.info(frame_info, track_message)
                    if gnb_id_search:
                        gnb_id = int(gnb_id_search.group(1))
                        frame_info = getframeinfo(currentframe())
                        track_message = f"gnb_id: {gnb_id}"
                        self.logger.info(frame_info, track_message)
                        gnb_folder_name_prefix_with_id = re.sub('\((id)\)', f'{gnb_id}', gnb_folder_name_prefix)
                        frame_info = getframeinfo(currentframe())
                        track_message = f"check gnb_dir.startswith(gnb_folder_name_prefix_with_id)"
                        self.logger.info(frame_info, track_message)
                        if gnb_dir.startswith(gnb_folder_name_prefix_with_id):
                            frame_info = getframeinfo(currentframe())
                            track_message = f"gnb_dir.startswith true"
                            self.logger.info(frame_info, track_message)
                            frame_info = getframeinfo(currentframe())
                            track_message = f"getting gnb_device_name"
                            self.logger.info(frame_info, track_message)
                            gnb_device_name = self.extract_gnb_name(gnb_dir, f'gnb_{gnb_id}_')
                            frame_info = getframeinfo(currentframe())
                            track_message = f"gnb_device_name: {gnb_device_name}"
                            self.logger.info(frame_info, track_message)
                            gnb_folder_index += 1
                            device_folder_index = 0
                            frame_info = getframeinfo(currentframe())
                            track_message = f"looping device_dir"
                            self.logger.info(frame_info, track_message)
                            for device_dir in listdir(os.path.join(sir_folder_path, gnb_dir)):
                                folder_name_prefix_regex = re.sub('.*\((id)\).*', '.*_([0-9]+)_.*', folder_name_prefix)
                                device_id_search = re.search(folder_name_prefix_regex, device_dir)
                                frame_info = getframeinfo(currentframe())
                                track_message = f"check is device_id_search"
                                self.logger.info(frame_info, track_message)
                                if device_id_search:
                                    frame_info = getframeinfo(currentframe())
                                    track_message = f"device_id_search true"
                                    self.logger.info(frame_info, track_message)
                                    device_id = int(device_id_search.group(1))
                                    folder_name_prefix_with_id = re.sub('\((id)\)', f'{device_id}', folder_name_prefix)
                                    frame_info = getframeinfo(currentframe())
                                    track_message = f"device_dir.startswith: {folder_name_prefix_with_id}"
                                    self.logger.info(frame_info, track_message)
                                    if device_dir.startswith(folder_name_prefix_with_id):
                                        frame_info = getframeinfo(currentframe())
                                        track_message = f"device_dir.startswith true"
                                        self.logger.info(frame_info, track_message)
                                        device_folder_index += 1
                                        frame_info = getframeinfo(currentframe())
                                        track_message = f"extract extract_device_name"
                                        self.logger.info(frame_info, track_message)
                                        device_name = self.extract_device_name(device_dir)
                                        frame_info = getframeinfo(currentframe())
                                        track_message = f"device_name: {device_name}"
                                        self.logger.info(frame_info, track_message)
                                        device_dir_path = os.path.join(sir_folder_path, gnb_dir, device_dir)
                                        device_dict = {'gnb_id': gnb_id,
                                                       'gnb_name': gnb_device_name,
                                                       'device_id': device_id,
                                                       'device_name': device_name,
                                                       'device_dir_path': device_dir_path}
                                        frame_info = getframeinfo(currentframe())
                                        track_message = f"set_filter: {set_filter}"
                                        self.logger.info(frame_info, track_message)
                                        if set_filter:
                                            if (set_filter['gnb_set_index'] is not None):
                                                if (set_filter['gnb_set_index'] == gnb_folder_index):
                                                    if (set_filter['device_set_index'] is not None):
                                                        if (set_filter['device_set_index'] == device_folder_index):
                                                            device_dict_list.append(device_dict)
                                                    else:
                                                        device_dict_list.append(device_dict)
                                            else:
                                                if (set_filter['device_set_index'] is not None):
                                                    if (set_filter['device_set_index'] == device_folder_index):
                                                        device_dict_list.append(device_dict)
                                                else:
                                                    device_dict_list.append(device_dict)
                                        else:
                                            if (gnb_folder_index == 1):
                                                device_dict_list.append(device_dict)
        return device_dict_list

    def append_log_dict_to_log_list(self, flow_type, last_log_file_path, device_name, flow_type_device_count, log_list):
        """
        Prepares a log dictionary and append it to the list of the logs' dictionary objects
        Arguments:
            flow_type: The flow type.
            last_log_file_path: The last log file for the specific log type and ru.
            device_name: The name of the device.
            flow_type_device_count: The number of rus.
            log_list: The list of the logs' dictionary objects.
            Returns:
                is_break: Breaks if a compatible regex_type has not been found for the log.
        """
        regex_type = self.preprocessing.check_regex_type(flow_type, last_log_file_path)
        if regex_type:
            log_dict = {
                'log_type': flow_type,
                'regex_type': regex_type,
                'log_file_path': last_log_file_path,
                'device_name': device_name,
                'flow_type_device_count': flow_type_device_count
            }
            log_list.append(log_dict)
            return False
        else:
            frame_info = getframeinfo(currentframe())
            track_message = f'a compatible regex_type has not been found for the log: {last_log_file_path} ' \
                            f'\n(file too short or not compatible pattern)'
            self.logger.error(frame_info, track_message)
            return True

    def prepare_specific_log_list(
            self,
            yaml_data,
            log_file_path,
            device_name):
        """
        Prepares a logsw dictionary list in Specific Log mode (When a given file is specified).
        Arguments:
            yaml_data: The YAML data object.
            log_file_path: The specified log file path.
            Returns:
                log_list: The list of the logs' dictionary objects.
                flow_type_device_count: The number of rus.
        """
        log_list = []
        flow_type_device_count = None
        if 'sequence_flows' in yaml_data:
            for sequence_flow_name in yaml_data['sequence_flows']:
                sequence_flow = yaml_data['sequence_flows'][sequence_flow_name]
                if 'flow_configuration' in sequence_flow:
                    sequence_flow_configuration = sequence_flow['flow_configuration']
                    if 'flow_type' in sequence_flow_configuration:
                        flow_type_device_count = 1
                        flow_type = sequence_flow_configuration['flow_type']
                        if not os.path.isfile(log_file_path):
                            frame_info = getframeinfo(currentframe())
                            track_message = 'The log file for the flow_type {flow_type} at path: {log_file_path} was not found.',
                            self.logger.error(frame_info, track_message)
                            break
                        is_break = self.append_log_dict_to_log_list(flow_type, log_file_path, device_name, flow_type_device_count, log_list)
                        if is_break:
                            break
        return log_list, flow_type_device_count

    def get_last_log_from_folder(
            self,
            device_id,
            device_name,
            device_dir_path,
            flow_type,
            is_log_analyzer_test):
        """
        Gets the last log file by datetime in a specific folder.
        Arguments:
            flow_type_folder_path: The path to the specific folder.
            flow_type: The flow type.
            is_log_analyzer_test: Weather the Log Analyzer is in Test Mode.
            Returns:
                last_log_file_path: The last log file for the specific log type and ru.
        """
        log_file_path = None
        last_log_file_path = None
        log_files = sorted(
            Path(device_dir_path).iterdir(),
            key=os.path.getmtime)[
                    ::-1]
        if is_log_analyzer_test:
            for log_file_path in log_files:
                if os.path.basename(log_file_path).startswith('Valid'):
                    last_log_file_path = log_file_path
                    break
        else:
            log_count = 0
            frame_info = getframeinfo(currentframe())
            track_message = f"looping log_files"
            self.logger.info(frame_info, track_message)
            for log_file_path in log_files:
                filename, file_extension = os.path.splitext(log_file_path)
                if (file_extension != '.bak'):
                    frame_info = getframeinfo(currentframe())
                    track_message = f"checking flow_type DU_MAC"
                    self.logger.info(frame_info, track_message)
                    if flow_type == 'DU_MAC':
                        frame_info = getframeinfo(currentframe())
                        track_message = f"flow_type is DU_MAC"
                        self.logger.info(frame_info, track_message)
                        frame_info = getframeinfo(currentframe())
                        track_message = f"checking _mac_ in path"
                        self.logger.info(frame_info, track_message)
                        if '_mac_' in os.path.basename(log_file_path):
                            frame_info = getframeinfo(currentframe())
                            track_message = f"_mac_ in path"
                            self.logger.info(frame_info, track_message)
                            frame_info = getframeinfo(currentframe())
                            track_message = f"checking log_count == 0"
                            self.logger.info(frame_info, track_message)
                            if log_count == 0:
                                frame_info = getframeinfo(currentframe())
                                track_message = f"log_count == 0"
                                self.logger.info(frame_info, track_message)
                                last_log_file_path = log_file_path
                                log_device_name = self.extract_device_name(os.path.basename(last_log_file_path))
                                frame_info = getframeinfo(currentframe())
                                track_message = f"log_device_name: {log_device_name}"
                                self.logger.info(frame_info, track_message)
                                if (log_device_name != '_'):
                                    frame_info = getframeinfo(currentframe())
                                    track_message = f"log_device_name: {log_device_name}"
                                    self.logger.info(frame_info, track_message)
                                    if log_device_name != device_name:
                                        return 'incompatible_folder_and_file_device_name'
                                    else:
                                        log_count += 1
                    frame_info = getframeinfo(currentframe())
                    track_message = f"is flow_type == 'DU_PHY'"
                    self.logger.info(frame_info, track_message)
                    if flow_type == 'DU_PHY' or flow_type == 'DU_PHY':
                        if '_phy_' in os.path.basename(log_file_path):
                            if log_count == 0:
                                last_log_file_path = log_file_path
                                log_device_name = self.extract_device_name(os.path.basename(last_log_file_path))
                                if (log_device_name != '_'):
                                    frame_info = getframeinfo(currentframe())
                                    track_message = f"log_device_name: {log_device_name}"
                                    self.logger.info(frame_info, track_message)
                                    if log_device_name != device_name:
                                        return 'incompatible_folder_and_file_device_name'
                                    else:
                                        log_count += 1
                    frame_info = getframeinfo(currentframe())
                    track_message = f"check flow_type == 'CUCP'"
                    self.logger.info(frame_info, track_message)
                    if flow_type == 'CUCP' or flow_type == 'CUUP':
                        frame_info = getframeinfo(currentframe())
                        track_message = f"flow_type == 'CUCP'"
                        self.logger.info(frame_info, track_message)
                        frame_info = getframeinfo(currentframe())
                        track_message = f"check describePods in path"
                        self.logger.info(frame_info, track_message)
                        frame_info = getframeinfo(currentframe())
                        track_message = f"describePods in path"
                        self.logger.info(frame_info, track_message)
                        frame_info = getframeinfo(currentframe())
                        track_message = f"check log_count == 0"
                        self.logger.info(frame_info, track_message)
                        if log_count == 0:
                            frame_info = getframeinfo(currentframe())
                            track_message = f"log_count == 0"
                            self.logger.info(frame_info, track_message)
                            last_log_file_path = log_file_path
                            log_device_name = self.extract_device_name(os.path.basename(last_log_file_path))
                            if (log_device_name != '_'):
                                frame_info = getframeinfo(currentframe())
                                track_message = f"log_device_name: {log_device_name}"
                                self.logger.info(frame_info, track_message)
                                if log_device_name != device_name:
                                    return 'incompatible_folder_and_file_device_name'
                                else:
                                    log_count += 1
            if log_count > 1:
                return 'multiple_logs'
            if not last_log_file_path:
                last_log_file_path = log_file_path
        frame_info = getframeinfo(currentframe())
        track_message = f"last_log_file_path: {last_log_file_path}"
        self.logger.info(frame_info, track_message)
        return last_log_file_path

    def extract_device_name(self, file_path):
        device_name_search = re.search('.*\((.+?)\).*', file_path)
        if device_name_search:
            return device_name_search.group(1)
        else:
            return '_'

    def extract_gnb_name(self, file_path, prefix):
        return file_path.split(prefix)[1]

    def extract_device_id(self, device_str, folder_name_prefix):
        folder_name_prefix_regex = re.sub('.*\((id)\).*', '\.\*\(id\)\.\*', folder_name_prefix)
        device_id = re.search(folder_name_prefix_regex, device_str).group(1)
        return device_id

    def validate_and_finalize(
            self,
            validation_data,
            yaml_data,
            sir_folder_path,
            flow_yaml_file_path,
            is_log_analyzer_test,
            is_create_yamls):
        """
        Runs the main validation and post validation functions.
        Arguments:
            validation_data: The validation data object.
            yaml_data: The YAML data object.
            sir_folder_path: the sir folder path.
            flow_yaml_file_path: The path to the YAML flow.
            is_log_analyzer_test: Weather the Log Analyzer is in Test Mode.
            Returns:
                post_validation_results: Post validation results.
        """
        validation_results = self.validation.validate(self.preprocessing, validation_data, yaml_data)
        global_validation_info = self.validation.process_global_actions(self.preprocessing, validation_data, yaml_data['global_configuration'])
        post_validation_results = self.post_validation(
            validation_results,
            validation_data,
            yaml_data,
            sir_folder_path,
            global_validation_info,
            flow_yaml_file_path,
            is_log_analyzer_test,
            is_create_yamls)
        return post_validation_results

    def analyze_log(
            self,
            sir_folder_path,
            flow_yaml_file_path,
            set_filter=None,
            preprocess_args=None,
            is_log_analyzer_test=None,
            results_yaml_path=None,
            is_create_yamls=True,
            under_or_supporting_test='under_test'):
        """
        Tests the analyze_log function and print the results
        Arguments:
            sir_folder_path: the sir folder path.
            flow_yaml_file_path: The path to the YAML flow.
            under_or_supporting_test: Weather the gnbs are under or supporting test
            preprocess_function_name: The preprocessing function to be called.
            preprocess_args: custom arguments for the preprocess function.
        """
        self.output.init_time_stats()
        self.output.create_results_folder(sir_folder_path)
        if is_create_yamls:
            self.output.store_flow_yaml(sir_folder_path, flow_yaml_file_path)
        validate_input_info = self.input_validation.validate_sir_folder(sir_folder_path)
        validate_input_info, validate_input_info_str = self.input_validation.validate_input_yaml(validate_input_info, flow_yaml_file_path)
        if validate_input_info['valid']:
            # self.output.copy_debug_mark_tests(sir_folder_path)
            if is_log_analyzer_test:
                log_list, flow_yaml_file_path, results_yaml_file_path, flow_type_device_count = \
                    self.prepare_input_files(sir_folder_path, is_log_analyzer_test, under_or_supporting_test, set_filter, flow_yaml_file_path)
            else:
                log_list, flow_yaml_file_path, flow_type_device_count = self.prepare_input_files(sir_folder_path,
                                                                                                 is_log_analyzer_test,
                                                                                                 under_or_supporting_test,
                                                                                                 set_filter,
                                                                                                 flow_yaml_file_path)
            if log_list == 'sed_missing':
                if is_log_analyzer_test:
                    return None, None, None, None, None
                else:
                    return None
            if log_list:
                self.init_analyzer(validate_input_info['yaml_data'])
                validate_input_info, validate_input_info_str = self.input_validation.validate_input_log(validate_input_info, log_list)
                if validate_input_info['valid']:
                    frame_info = getframeinfo(currentframe())
                    self.logger.info(frame_info, validate_input_info_str)
                    validation_data, validation_type, yaml_data = self.preprocessing.preprocess(log_list, validate_input_info, self.validation.validation_state, preprocess_args)
                    if validation_data:
                        results = self.validate_and_finalize(validation_data, yaml_data, sir_folder_path, flow_yaml_file_path, is_log_analyzer_test, is_create_yamls)
                        if is_log_analyzer_test:
                            results_valid, original_valid, original_end_reason = self.validate_results(results, results_yaml_path)
                            return results_valid, results['valid'], results['validation_end_reason'], original_valid, original_end_reason
                        else:
                            return results
                else:
                    frame_info = getframeinfo(currentframe())
                    self.logger.error(frame_info, validate_input_info_str)
            else:
                results = {
                    'valid': 'FAIL',
                    'validation_end_reason': 'log list preparation failed.',
                    'time_stats': self.output.time_stats}
                self.output.finalize_time_stats()
                self.output.output_validation_results_log_list_error(sir_folder_path, results, flow_yaml_file_path, is_log_analyzer_test)
                return results
        else:
            track_message = ""
            for fail_tests in validate_input_info['fail_tests_list']:
                track_message += fail_tests['test_description']
            frame_info = getframeinfo(currentframe())
            self.logger.error(frame_info, track_message)

    def analyze_specific_logs(
            self,
            log_file_path,
            flow_yaml_file_path,
            preprocess_args=None,
    ):
        """
        Tests the analyze_log function and print the results
        Arguments:
            log_file_path: The path to the log file.
            flow_yaml_file_path: The path to the YAML flow.
            preprocess_function_name: The preprocessing function to be called.
            preprocess_args: custom arguments for the preprocess function.
        Returns:
            results: The results object.
        """
        self.output.init_time_stats()
        is_log_analyzer_test = False
        results = None
        device_name = 'unspecified'
        log_list, flow_yaml_file_path, flow_type_device_count = self.prepare_specific_input_files(log_file_path, device_name, is_log_analyzer_test, flow_yaml_file_path)
        if log_list:
            validate_input_info, validate_input_info_str = self.input_validation.validate_input_yaml(flow_yaml_file_path)
            self.init_analyzer(validate_input_info['yaml_data'])
            validate_input_info, validate_input_info_str = self.input_validation.validate_input_log(
                validate_input_info, log_list)
            if validate_input_info['valid']:
                frame_info = getframeinfo(currentframe())
                self.logger.info(frame_info, validate_input_info_str)
                validation_data, validation_type, yaml_data = self.preprocessing.preprocess(
                    log_list, validate_input_info, self.validation.validation_state, preprocess_args)
                if validation_data:
                    results = self.validate_and_finalize(
                        validation_data, yaml_data, None, flow_yaml_file_path, is_log_analyzer_test)
                    self.output.finalize_time_stats()
                    return results
            else:
                self.output.finalize_time_stats()
                results = {'valid': 'FAIL',
                           'validation_end_reason': 'Input validation failed',
                           'time_stats': self.output.time_stats}
                frame_info = getframeinfo(currentframe())
                self.logger.error(frame_info, validate_input_info_str)
        else:
            self.output.finalize_time_stats()
            results = {
                'valid': 'FAIL',
                'validation_end_reason': 'One or more files was too short or their regex pattern was not found.',
                'time_stats': self.output.time_stats}
        return results

    def test_log_analyzer(self, under_or_supporting_test='under_test'):
        """
        Loops over all the test in the excel file and update the results.
        """
        if not self.test_network_available():
            frame_info = getframeinfo(currentframe())
            track_message = f"The network is not available"
            self.logger.error(frame_info, track_message)
        else:
            test_excel_path = os.path.join(
                os.path.dirname(__file__),
                'Tests',
                'tests_xlsx',
                'log_analyzer_tests.xlsx')
            read_tests_workbook = openpyxl.load_workbook(
                test_excel_path, read_only=True, data_only=True)
            write_tests_workbook = openpyxl.load_workbook(test_excel_path)
            read_worksheet = read_tests_workbook["Tests"]
            write_worksheet = write_tests_workbook["Tests"]
            cols_name = {}
            curr_col = 0
            for col in write_worksheet.iter_cols(
                    1, write_worksheet.max_column):
                cols_name[col[1].value] = curr_col
                curr_col += 1

            single_tests_count = 0
            test_number = None
            hostname = socket.gethostname()
            if hostname.upper() == 'ASIL-RBLUMBERG1':
                test_folder_path = r'C:\Users\Administrator\main\projects\LogAnalyzer_Tests'
            elif hostname.upper() == 'ASIL-RBLUMBERG':
                test_folder_path = r'C:\Users\rblumberg\main\projects\LogAnalyzer_Tests'
            elif hostname.upper() == 'ASIL-EZLIFE-DEV':
                test_folder_path = r'C:\Users\Administrator\PycharmProjects\LogAnalyzer_Tests'
            else:
                test_folder_path = None
            for row in read_worksheet.iter_rows(min_row=3, max_row=500):
                enabled = row[cols_name['enabled']].value
                if enabled == 'single_test':
                    single_tests_count += 1
                    test_number = row[cols_name['#']].row - 2
            if single_tests_count == 0:
                for row in read_worksheet.iter_rows(min_row=3, max_row=500):
                    enabled = row[cols_name['enabled']].value
                    if enabled:
                        test_number = row[cols_name['#']].row - 2
                        if test_number is None:
                            break
                        else:
                            row_num = 2 + test_number
                        if test_folder_path:
                            sir_folder_rel_path = row[cols_name['test folder path']].value
                            sir_folder_path = os.path.join(test_folder_path, sir_folder_rel_path)
                            if sir_folder_path and os.path.isdir(sir_folder_path):
                                sir_folder_name = os.path.basename(sir_folder_path)
                                frame_info = getframeinfo(currentframe())
                                track_message = f'start of test: {sir_folder_name}'
                                self.logger.info(frame_info, track_message)
                                self.update_info_in_excel(sir_folder_name, write_worksheet, row_num, cols_name)
                                is_log_analyzer_test = True
                                arguments = row[cols_name['arguments']].value
                                if arguments is not None:
                                    arguments_dict = json.loads(arguments)
                                else:
                                    arguments_dict = None
                                if arguments_dict and 'set_filter' in arguments_dict:
                                    log_list, flow_yaml_file_path, results_yaml_file_path, flow_type_device_count = \
                                        self.prepare_input_files(sir_folder_path, is_log_analyzer_test, under_or_supporting_test, arguments_dict['set_filter'])
                                else:
                                    log_list, flow_yaml_file_path, results_yaml_file_path, flow_type_device_count = \
                                        self.prepare_input_files(sir_folder_path, is_log_analyzer_test, under_or_supporting_test)
                                if (not log_list) or (not flow_yaml_file_path) or (not results_yaml_file_path):
                                    frame_info = getframeinfo(currentframe())
                                    track_message = "prepare test failed"
                                    self.logger.error(frame_info, track_message)
                                else:
                                    if log_list != 'sed_missing':
                                        self.validate_and_update(sir_folder_path, flow_yaml_file_path, results_yaml_file_path, write_worksheet, row_num, cols_name)
                                    else:
                                        break

                            else:
                                frame_info = getframeinfo(currentframe())
                                track_message = "Test folder doesn't exist"
                                self.logger.error(frame_info, track_message)
                        else:
                            frame_info = getframeinfo(currentframe())
                            track_message = "Hostname not recognized and test folder not defined."
                            self.logger.error(frame_info, track_message)
            elif single_tests_count == 1:
                row_num = 2 + test_number
                if test_folder_path:
                    sir_folder_rel_path = read_worksheet[row_num][cols_name['test folder path']].value
                    sir_folder_path = os.path.join(test_folder_path, sir_folder_rel_path)
                    if sir_folder_path and os.path.isdir(sir_folder_path):
                        sir_folder_name = os.path.basename(sir_folder_path)
                        self.update_info_in_excel(sir_folder_name, write_worksheet, row_num, cols_name)
                        is_log_analyzer_test = True
                        arguments = read_worksheet[row_num][cols_name['arguments']].value
                        if arguments is not None:
                            arguments_dict = json.loads(arguments)
                        else:
                            arguments_dict = None
                        if arguments_dict and 'set_filter' in arguments_dict:
                            log_list, flow_yaml_file_path, results_yaml_file_path, flow_type_device_count = \
                                self.prepare_input_files(sir_folder_path, is_log_analyzer_test, under_or_supporting_test, arguments_dict['set_filter'])
                        else:
                            log_list, flow_yaml_file_path, results_yaml_file_path, flow_type_device_count = \
                                self.prepare_input_files(sir_folder_path, is_log_analyzer_test, under_or_supporting_test)
                        if (not log_list) or (not flow_yaml_file_path) or (not results_yaml_file_path):
                            frame_info = getframeinfo(currentframe())
                            track_message = "prepare test failed"
                            self.logger.error(frame_info, track_message)
                        elif not log_list:
                            results = {
                                'valid': 'FAIL',
                                'validation_end_reason': 'One or more files was too short or their regex pattern was not found.',
                                'time_stats': self.output.time_stats}
                            results_valid, original_valid, original_end_reason = self.validate_results(results, results_yaml_file_path)
                            self.update_validation_results_in_excel(results_valid, write_worksheet[row_num][cols_name['valid']])
                            self.update_original_valid_result_in_excel(original_valid, write_worksheet[row_num][cols_name['Original Result']])
                            self.update_analysis_valid_result_in_excel('FAIL', write_worksheet[row_num][cols_name['Analysis Result']])
                            self.update_original_end_reason_in_excel(original_end_reason, write_worksheet[row_num][cols_name['original end reason']])
                            self.update_analysis_end_reason_in_excel(results['validation_end_reason'], write_worksheet[row_num][cols_name['received end reason']])
                        else:
                            if arguments_dict and 'set_filter' in arguments_dict:
                                self.validate_and_update(sir_folder_path,
                                                         flow_yaml_file_path,
                                                         results_yaml_file_path,
                                                         write_worksheet,
                                                         row_num,
                                                         cols_name,
                                                         arguments_dict['set_filter'])
                            else:
                                self.validate_and_update(sir_folder_path,
                                                         flow_yaml_file_path,
                                                         results_yaml_file_path,
                                                         write_worksheet,
                                                         row_num,
                                                         cols_name)
                    else:
                        frame_info = getframeinfo(currentframe())
                        track_message = "Test folder doesn't exist"
                        self.logger.error(frame_info, track_message)
                else:
                    frame_info = getframeinfo(currentframe())
                    track_message = "Hostname not recognized and test folder not defined."
                    self.logger.error(frame_info, track_message)
            elif single_tests_count > 1:
                frame_info = getframeinfo(currentframe())
                track_message = "more than one single test has been specified"
                self.logger.error(frame_info, track_message)
            try:
                write_tests_workbook.save(test_excel_path)
            except Exception:
                frame_info = getframeinfo(currentframe())
                track_message = "Can't save excel spreadsheet (it is probably opened)."
                self.logger.error(frame_info, track_message)

    def update_info_in_excel(self, sir_folder_name, write_worksheet, row_num, cols_name):
        """
        Update SIR and test name in Excel
        """
        self.update_test_name_in_excel(sir_folder_name, write_worksheet[row_num][cols_name['test name']])

    def validate_and_update(self, sir_folder_path, flow_yaml_file_path, results_yaml_file_path, write_worksheet, row_num, cols_name, set_filter=None):
        """
        For each test in the excel do an analysis test and update the results.
        Arguments:
            sir_folder_path: the sir folder path.
            results_yaml_file_path: The path the Results YAML file.
            write_worksheet: The excel worksheet.
            row_num: The row number of the test.
            cols_name: The column number of the field.
        """
        preprocess_args = None
        results_valid, analysis_valid, analysis_end_reason, original_valid, original_end_reason = \
            self.analyze_log(sir_folder_path=sir_folder_path,
                             flow_yaml_file_path=flow_yaml_file_path,
                             preprocess_args=preprocess_args,
                             is_log_analyzer_test=True,
                             results_yaml_path=results_yaml_file_path,
                             set_filter=set_filter)
        self.update_validation_results_in_excel(results_valid, write_worksheet[row_num][cols_name['valid']])
        self.update_original_valid_result_in_excel(original_valid, write_worksheet[row_num][cols_name['Original Result']])
        self.update_analysis_valid_result_in_excel(analysis_valid, write_worksheet[row_num][cols_name['Analysis Result']])
        self.update_original_end_reason_in_excel(original_end_reason, write_worksheet[row_num][cols_name['original end reason']])
        self.update_analysis_end_reason_in_excel(analysis_end_reason, write_worksheet[row_num][cols_name['received end reason']])

    def post_validation(self,
                        results,
                        validation_data,
                        yaml_data,
                        sir_folder_path,
                        global_validation_info,
                        flow_yaml_file_path,
                        is_log_analyzer_test,
                        is_create_yamls):
        """
        Creates the final results dict and output the validation results,
        Arguments:
            results: The results object. 
            validation_data: The validation data object.
            yaml_data: The YAML data object.
            sir_folder_path: the SIR folder path.
            global_validation_info: The global validation information.
            flow_yaml_file_path: The path to the YAML flow.
            is_log_analyzer_test: Weather the Log Analyzer is in Test Mode.
        Returns:
            results: The final results object.
        """
        self.output.finalize_time_stats()
        results = self.output.create_results_dict(results, global_validation_info, validation_data, yaml_data)
        self.output.output_validation_results(sir_folder_path,
                                              results,
                                              flow_yaml_file_path,
                                              is_log_analyzer_test,
                                              is_create_yamls)
        if not is_log_analyzer_test:
            self.send_push_notification(results)
        return results

    @staticmethod
    def update_validation_results_in_excel(results_valid, result_cell):
        """
        Updates the validation result for a specific test in the Excel file.
        Arguments:
            results_valid: Weather the result is valid. 
            result_cell: The cell to be updated.
        """
        if results_valid:
            result_cell.value = 'PASS'
        else:
            result_cell.value = 'FAIL'

    @staticmethod
    def update_test_name_in_excel(test_name, test_name_cell):
        """
        Updates the test name for the test in the Excel file.
        Arguments:
            test_name: the test name. 
            test_name_cell: The cell to be updated.
        """
        test_name_cell.value = test_name

    @staticmethod
    def update_original_valid_result_in_excel(
            original_valid_result,
            original_valid_result_cell):
        """
        Updates the original validation value (Taken from the Results YAML) in the Excel file.
        Arguments:
            original_valid_result: The original validation result.
            original_valid_result_cell: The cell to be updated.
        """
        original_valid_result_cell.value = original_valid_result

    @staticmethod
    def update_analysis_valid_result_in_excel(analysis_valid_result, analysis_valid_result_cell):
        """
        Updates the analysis validation value in the Excel file.
        Arguments:
            analysis_valid_result: The analysis validation result.
            analysis_valid_result_cell: The cell to be updated.
        """
        analysis_valid_result_cell.value = analysis_valid_result

    @staticmethod
    def update_original_end_reason_in_excel(end_reason, end_reason_cell):
        """
        Updates the original end reason value in the Excel file.
        Arguments:
            end_reason: The original end reason.
            end_reason_cell: The cell to be updated.
        """
        end_reason_cell.value = end_reason

    @staticmethod
    def update_analysis_end_reason_in_excel(end_reason, end_reason_cell):
        """
        Updates the analysis end reason value in the Excel file.
        Arguments:
            end_reason: The analysis end reason.
            end_reason_cell: The cell to be updated.
        """
        end_reason_cell.value = end_reason

    @staticmethod
    def validate_results(results, results_yaml_path):
        """
        Checks weather the analysis validation results are in alignment with the original validation results.
        Arguments:
            results: The final results object of the analysis.
            results_yaml_path: The path to the original results YAML.
        Returns:
            results_valid: Weather the validation results are in alignment.
            original_valid: The original validation results.
            original_end_reason: The original end reason.
        """
        yaml_str = open(results_yaml_path, 'r')
        original_results = yaml.load(yaml_str, Loader=yaml.Loader)
        if (results['valid'] == original_results['valid']) and (results['validation_end_reason'] == original_results['validation_end_reason']):
            if 'validated_flows' in results:
                for results_validated_flow_index, results_validated_flow in enumerate(results['validated_flows']):
                    results_validated_flow_name = list(original_results['validated_flows'])[results_validated_flow_index]
                    validated_flow = results['validated_flows'][results_validated_flow_name]
                    original_validated_flow = original_results['validated_flows'][results_validated_flow_name]
                    if (validated_flow['sequence_flow_validation_result'] != original_validated_flow['sequence_flow_validation_result']) and (
                            validated_flow['sequence_flow_validation_end_reason'] != original_validated_flow['sequence_flow_validation_end_reason']):
                        return False, original_results['valid'], original_results['validation_end_reason']
                return True, original_results['valid'], original_results['validation_end_reason']
            else:
                return True, original_results['valid'], original_results['validation_end_reason']
        else:
            return False, original_results['valid'], original_results['validation_end_reason']

    def prepare_input_files(
            self,
            sir_folder_path,
            is_log_analyzer_test,
            under_or_supporting_test,
            set_filter=None,
            flow_yaml_file_path=None):
        """
        Prepares the input files
        Arguments:
            sir_folder_path: the SIR folder path.
            is_log_analyzer_test: Weather the Log Analyzer is in Test Mode.
            flow_yaml_file_path: The path to the YAML flow.
        Returns:
            log_list: The list of the logs' dictionary objects.
            flow_yaml_file_path: The path to the YAML flow.
            flow_type_device_count: The number of rus.
        """
        flow_yaml_file_path, yaml_data = self.prepare_yaml_flow(sir_folder_path, is_log_analyzer_test, flow_yaml_file_path)
        if is_log_analyzer_test:
            if not yaml_data:
                return None, None, None, None
            log_list, flow_type_device_count = self.prepare_log_list(yaml_data, sir_folder_path, is_log_analyzer_test, under_or_supporting_test, set_filter)
            results_yaml_file_path = self.prepare_results_yaml(sir_folder_path)
            return log_list, flow_yaml_file_path, results_yaml_file_path, flow_type_device_count
        else:
            if not yaml_data:
                return None, None, None
            log_list, flow_type_device_count = self.prepare_log_list(yaml_data, sir_folder_path, is_log_analyzer_test, under_or_supporting_test, set_filter)
        return log_list, flow_yaml_file_path, flow_type_device_count

    def prepare_specific_input_files(
            self,
            log_file_path,
            device_name,
            is_log_analyzer_test,
            flow_yaml_file_path=None):
        """
        Prepares the input files in Specific Log mode
        Arguments:
            log_file_path: The path to the log file.
            is_log_analyzer_test: Weather the Log Analyzer is in Test Mode.
            flow_yaml_file_path: The path to the YAML flow.
        Returns:
            log_list: The list of the logs' dictionary objects.
            flow_yaml_file_path: The path to the YAML flow.
            flow_type_device_count: The number of rus.
        """
        flow_yaml_file_path, yaml_data = self.prepare_yaml_flow(None, is_log_analyzer_test, flow_yaml_file_path)
        if not yaml_data:
            return None, None, None
        log_list, flow_type_device_count = self.prepare_specific_log_list(yaml_data, log_file_path, device_name)
        return log_list, flow_yaml_file_path, flow_type_device_count

    def prepare_results_yaml(self, test_folder_path):
        """
        Prepares the results YAML.
        Arguments:
            test_folder_path: The path to the test (SIR) folder.
        Returns:
            valid_results_yaml_file_path: The path to the results YAML or none of none exists.
        """
        results_yaml_folder_path = os.path.join(
            test_folder_path, 'LogAnalyzerResults')
        results_yaml_files_paths = sorted(
            Path(results_yaml_folder_path).iterdir(),
            key=os.path.getmtime)[
                                   ::-1]
        valid_results_yaml_file_path = None
        for results_yaml_file_path in results_yaml_files_paths:
            if os.path.basename(results_yaml_file_path).startswith('Valid'):
                valid_results_yaml_file_path = results_yaml_file_path
                break
        if not valid_results_yaml_file_path:
            frame_info = getframeinfo(currentframe())
            track_message = "No valid Results YAML flow file was found in YAML flow folder path."
            self.logger.error(frame_info, track_message)
        return valid_results_yaml_file_path

    def prepare_yaml_flow(
            self,
            test_folder_path,
            is_log_analyzer_test,
            flow_yaml_file_path=None):
        """
        Prepares the flow YAML.
        Arguments:
            test_folder_path: The path to the test (SIR) folder.
            is_log_analyzer_test: Weather the Log Analyzer is in Test Mode.
             flow_yaml_file_path: The path to the YAML flow.
        Returns:
            flow_yaml_file_path: The path to the YAML flow.
            yaml_data: The YAML data object.
        """
        if is_log_analyzer_test:
            yaml_flow_folder_path = os.path.join(test_folder_path, 'YAML_flow')
            yaml_flow_folder_paths = sorted(
                Path(yaml_flow_folder_path).iterdir(),
                key=os.path.getmtime)[
                                     ::-1]
            if is_log_analyzer_test:
                valid_flow_yaml_file_path = None
                for flow_yaml_file_path in yaml_flow_folder_paths:
                    if os.path.basename(
                            flow_yaml_file_path).startswith('Valid'):
                        valid_flow_yaml_file_path = flow_yaml_file_path
                        break
            else:
                valid_flow_yaml_file_path = yaml_flow_folder_paths[0]
            if not valid_flow_yaml_file_path:
                frame_info = getframeinfo(currentframe())
                track_message = "No valid YAML flow file was found in YAML flow folder path."
                self.logger.error(frame_info, track_message)
                return None, None
            yaml_data = self.input_validation.load_yaml_from_file(
                valid_flow_yaml_file_path)
            return flow_yaml_file_path, yaml_data
        else:
            yaml_data = self.input_validation.load_yaml_from_file(
                flow_yaml_file_path)
            return flow_yaml_file_path, yaml_data

    @staticmethod
    def create_validation_summary(results):
        """
        Creates the validation summary.
        Arguments:
            results: The final results object of the analysis.
        Returns:
            validation_summary: The validation summary.
        """
        flows_name_list = []
        for flow in results['validated_flows']:
            flows_name_list.append(flow)
        flows_str = ','.join(flows_name_list)
        validation_summary = f"The flows: {flows_str} validated, " \
                             f"The run result is: {results['valid']}, the end reason is " \
                             f"{results['validation_end_reason']}"
        return validation_summary

    def create_validation_summary_error(self, results):
        """
        Creates the validation summary when error occurred.
        Arguments:
            results: The final results object of the analysis.
        Returns:
            validation_summary: The validation summary.
        """
        flows_name_list = []
        for flow in results['validated_flows']:
            flows_name_list.append(flow)
        errors_str = repr(self.logger.logger_lists['error_list'])
        validation_summary = f"A run ended with error.\n details:\n{errors_str}"
        return validation_summary

    # @staticmethod
    # def get_folder_sub_dirs(folder_path):
    #     result = []
    #     for subdir in os.listdir(folder_path):
    #         subdir_path = os.path.join(folder_path, subdir)
    #         if os.path.isdir(subdir_path):
    #             result.append(subdir_path)
    #     return result

    def send_push_notification(self, results):
        """
        Send a push notification with a validation summary.
        Arguments:
            results: The final results object of the analysis.
        """
        hostname = os.getenv('COMPUTERNAME')
        if hostname != 'ASIL-RBLUMBERG' and hostname != 'ASIL-RBLUMBERG1':
            if self.logger.logger_lists['error_list']:
                validation_summary = self.create_validation_summary_error(
                    results)
                self.notifications.notify(
                    'A Log Analyzer run has ended with error:',
                    validation_summary)
            else:
                validation_summary = self.create_validation_summary(results)
                self.notifications.notify(
                    'A Log Analyzer run has ended:', validation_summary)

    @staticmethod
    def test_network_available():
        """
        Test weather the network is available (Requires VPN connection if outside of office).
        """
        network_folder_path = r'\\192.168.127.231\LabShare\Automation\LogAnalyzer\debug'
        if os.path.isdir(network_folder_path):
            return True
        else:
            return False

    @staticmethod
    def configure_locale():
        """
        Configures the locale (Solved a bug during a previous run).
        """
        locale.setlocale(locale.LC_ALL, 'en_US.utf8')
